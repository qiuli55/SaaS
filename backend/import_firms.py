"""万能律所导入脚本 -- 把从各省市开放数据平台下载的 Excel/CSV 扔进来"""
import sqlite3, os, sys, csv
from pathlib import Path

DB = os.path.join(os.path.dirname(__file__), "legal_ai.db")


def import_xlsx(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return 0
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return _import_rows(rows[1:], rows[0])  # skip header


def import_csv(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return _import_rows(rows[1:], rows[0])


def _import_rows(rows, header):
    """从表格行中提取��所名称"""
    conn = sqlite3.connect(DB)
    conn.execute("""CREATE TABLE IF NOT EXISTS law_firms (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name VARCHAR(200), city VARCHAR(50),
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # 查找名称列（自动匹配列名）
    name_idx = city_idx = None
    for i, h in enumerate(header):
        h = str(h).strip()
        if any(k in h for k in ["律所名称", "律师事务所名称", "事务所名称", "机构名称", "中文名称", "名称", "name", "firm"]):
            name_idx = i
        if any(k in h for k in ["城市", "所在地区", "区", "所属城市", "city", "province", "所属区县"]):
            city_idx = i

    if name_idx is None:
        print(f"  警告: 无法自动找到律所名称列, header: {list(header)[:8]}")
        return 0

    count = 0
    seen = set()
    for row in rows:
        if not row or len(row) <= name_idx:
            continue
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        if len(name) < 6 or "律师事务所" not in name:
            continue
        if name in seen:
            continue
        seen.add(name)
        city = str(row[city_idx]).strip() if city_idx is not None and len(row) > city_idx else ""
        conn.execute(
            "INSERT OR IGNORE INTO law_firms (name, city) VALUES (?,?)",
            (name, city),
        )
        count += 1

    conn.commit()
    conn.close()
    return count


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python import_firms.py 文件1.xlsx 文件2.csv ...")
        print("示例: python import_firms.py 律师事务所信息.xls 中山市律所.csv")
        sys.exit(1)

    total = 0
    for filepath in sys.argv[1:]:
        if not os.path.exists(filepath):
            print(f"跳过: {filepath} 不存在")
            continue
        print(f"处理 {filepath}...")
        ext = filepath.lower()
        if ext.endswith((".xlsx", ".xls")):
            count = import_xlsx(filepath)
        elif ext.endswith(".csv"):
            count = import_csv(filepath)
        else:
            print(f"  不支持的文件格式: {ext}")
            continue
        print(f"  导入 {count} 条")
        total += count

    # 显示总计
    conn = sqlite3.connect(DB)
    final = conn.execute("SELECT COUNT(*) FROM law_firms").fetchone()[0]
    conn.close()
    print(f"\n=== 完成 === 本次导入 {total} 条, 数据库总计 {final} 条")
