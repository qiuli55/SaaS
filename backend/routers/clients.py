from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from database import get_db
from models import Client, User, Case
import io
import json
from schemas import ClientCreate, ClientUpdate
from auth import get_current_user

router = APIRouter(prefix="/api/clients", tags=["客户"])


@router.get("")
def list_clients(
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Client).filter(Client.user_id == current_user.id)
    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(
            (Client.name.contains(kw)) |
            (Client.phone.contains(kw)) |
            (Client.company.contains(kw))
        )

    total = query.count()
    clients = query.order_by(Client.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    items = []
    for c in clients:
        case_count = db.query(Case).filter(Case.client_id == c.id).count()
        items.append({
            "id": c.id,
            "user_id": c.user_id,
            "name": c.name,
            "phone": c.phone,
            "wechat": c.wechat,
            "id_card": c.id_card,
            "company": c.company,
            "tags": c.tags,
            "remark": c.remark,
            "created_at": c.created_at.isoformat() if c.created_at else "",
            "case_count": case_count,
        })

    return {"code": 0, "data": {"total": total, "items": items, "page": page, "page_size": page_size}}


@router.post("")
def create_client(
    req: ClientCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = Client(
        user_id=current_user.id,
        name=req.name or "",
        phone=req.phone or "",
        wechat=req.wechat or "",
        id_card=req.id_card or "",
        company=req.company or "",
        tags=req.tags or "",
        remark=req.remark or "",
    )
    db.add(client)
    db.commit()
    db.refresh(client)

    return {
        "code": 0,
        "message": "创建成功",
        "data": {
            "id": client.id,
            "user_id": client.user_id,
            "name": client.name,
            "phone": client.phone,
            "wechat": client.wechat,
            "id_card": client.id_card,
            "company": client.company,
            "tags": client.tags,
            "remark": client.remark,
            "created_at": client.created_at.isoformat() if client.created_at else "",
            "case_count": 0,
        },
    }


@router.get("/export")
def export_clients(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出客户为 Excel"""
    clients = db.query(Client).filter(Client.user_id == current_user.id).order_by(Client.created_at.desc()).all()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "客户通讯录"

    headers = ["姓名", "手机号", "微信", "身份证号", "公司", "标签", "备注", "创建时间"]
    ws.append(headers)

    for c in clients:
        tags = c.tags or ""
        try:
            parsed = json.loads(tags)
            tags = ", ".join(parsed) if isinstance(parsed, list) else str(parsed)
        except:
            pass
        created = ""
        try:
            if c.created_at:
                created = c.created_at.strftime("%Y-%m-%d %H:%M")
        except:
            pass
        ws.append([c.name or "", c.phone or "", c.wechat or "", c.id_card or "",
                   c.company or "", tags, c.remark or "", created])

    for col, w in zip("ABCDEFGH", [12, 14, 14, 20, 20, 20, 30, 18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=clients.xlsx"})


@router.post("/import")
async def import_clients(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从 Excel 批量导入客户"""
    from openpyxl import load_workbook
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip() if row[0] else None
        if not name:
            continue
        phone = str(row[1]).strip() if len(row) > 1 and row[1] else None
        wechat = str(row[2]).strip() if len(row) > 2 and row[2] else None
        id_card = str(row[3]).strip() if len(row) > 3 and row[3] else None
        company = str(row[4]).strip() if len(row) > 4 and row[4] else None
        tags = str(row[5]).strip() if len(row) > 5 and row[5] else None
        remark = str(row[6]).strip() if len(row) > 6 and row[6] else None

        client = Client(
            user_id=current_user.id, name=name, phone=phone, wechat=wechat,
            id_card=id_card, company=company, tags=tags, remark=remark
        )
        db.add(client)
        count += 1
    db.commit()
    return {"message": f"成功导入 {count} 位客户", "count": count}


@router.get("/{client_id}")
def get_client(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = db.query(Client).filter(
        Client.id == client_id, Client.user_id == current_user.id
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="客户不存在")

    # 获取关联案件
    cases = db.query(Case).filter(Case.client_id == client_id).all()
    case_list = []
    for c in cases:
        case_list.append({
            "id": c.id,
            "case_no": c.case_no,
            "case_type": c.case_type,
            "plaintiff": c.plaintiff,
            "defendant": c.defendant,
            "status": c.status,
            "subject_amount": float(c.subject_amount or 0),
        })

    return {
        "code": 0,
        "data": {
            "id": client.id,
            "user_id": client.user_id,
            "name": client.name,
            "phone": client.phone,
            "wechat": client.wechat,
            "id_card": client.id_card,
            "company": client.company,
            "tags": client.tags,
            "remark": client.remark,
            "created_at": client.created_at.isoformat() if client.created_at else "",
            "case_count": len(case_list),
            "cases": case_list,
        },
    }


@router.put("/{client_id}")
def update_client(
    client_id: int,
    req: ClientUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = db.query(Client).filter(
        Client.id == client_id, Client.user_id == current_user.id
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="客户不存在")

    update_data = req.model_dump(exclude_unset=True)
    for k, v in update_data.items():
        setattr(client, k, v)

    db.commit()
    db.refresh(client)

    return {
        "code": 0,
        "message": "更新成功",
        "data": {
            "id": client.id,
            "name": client.name,
            "phone": client.phone,
            "wechat": client.wechat,
            "id_card": client.id_card,
            "company": client.company,
            "tags": client.tags,
            "remark": client.remark,
        },
    }


@router.delete("/{client_id}")
def delete_client(
    client_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    client = db.query(Client).filter(
        Client.id == client_id, Client.user_id == current_user.id
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="客户不存在")
    db.delete(client)
    db.commit()
